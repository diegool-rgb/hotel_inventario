from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import ensure_csrf_cookie
from django.db.models import Sum, Count, Q
from django.http import JsonResponse
from django.contrib import messages
from django.db import transaction
from django.urls import reverse
from django.utils.http import urlencode
from decimal import Decimal
from datetime import date
from .models import Producto, Stock, Movimiento, AlertaStock, Area, Categoria, Proveedor, EntradaStock, DetalleEntradaStock
from .forms import AgregarProductoForm, EntradaStockForm, DetalleEntradaFormSet, ProveedorForm
import json
import time
from django.http import HttpResponse
import csv


def home(request):
    """Vista de inicio - página principal"""
    if request.user.is_authenticated:
        # Si está autenticado, redirigir al dashboard
        from django.shortcuts import redirect
        return redirect('inventario:dashboard')
    else:
        # Si no está autenticado, redirigir al login
        from django.shortcuts import redirect
        return redirect('usuarios:login')


@login_required
def dashboard(request):
    """Vista principal del dashboard"""
    from django.utils import timezone
    from datetime import timedelta
    
    # Estadísticas generales
    total_productos = Producto.objects.filter(activo=True).count()
    total_categorias = Categoria.objects.filter(activo=True).count()
    
    # Calcular productos con stock bajo
    productos_bajo_stock = []
    productos_con_stock = Producto.objects.filter(activo=True).prefetch_related('stocks')
    
    for producto in productos_con_stock:
        stock_total = producto.stock_total()
        if stock_total <= producto.stock_minimo:
            productos_bajo_stock.append(producto)
    
    # Movimientos recientes (últimos 7 días o últimos 10)
    fecha_limite = timezone.now() - timedelta(days=7)
    movimientos_recientes = Movimiento.objects.select_related(
        'producto', 'usuario'
    ).filter(
        fecha__gte=fecha_limite
    ).order_by('-fecha')[:10]
    
    # Si no hay movimientos recientes, tomar los últimos 10
    if not movimientos_recientes:
        movimientos_recientes = Movimiento.objects.select_related(
            'producto', 'usuario'
        ).order_by('-fecha')[:10]
    
    # Movimientos del día de hoy
    hoy = timezone.now().date()
    movimientos_hoy = Movimiento.objects.filter(
        fecha__date=hoy
    ).count()
    
    # Productos por categoría
    productos_por_categoria = Categoria.objects.filter(activo=True).annotate(
        total=Count('productos', filter=Q(productos__activo=True))
    ).order_by('-total')[:5]
    
    # Calcular porcentajes para las categorías
    for categoria in productos_por_categoria:
        if total_productos > 0:
            categoria.porcentaje = (categoria.total * 100) / total_productos
        else:
            categoria.porcentaje = 0
    
    # Valor total del inventario (solo productos con precio)
    valor_inventario = Decimal('0')
    for producto in productos_con_stock:
        if producto.precio_unitario:
            valor_inventario += producto.stock_total() * producto.precio_unitario
    
    # Preparar estadísticas para el template
    stats = {
        'total_productos': total_productos,
        'productos_bajo_stock': len(productos_bajo_stock),
        'movimientos_hoy': movimientos_hoy,
        'valor_inventario': valor_inventario,
    }

    context = {
        'stats': stats,
        'productos_bajo_stock': productos_bajo_stock[:10],  # Solo mostrar 10
        'movimientos_recientes': movimientos_recientes,
        'productos_por_categoria': productos_por_categoria,
        'total_productos': total_productos,
        'total_categorias': total_categorias,
    }

    return render(request, 'inventario/dashboard.html', context)


@login_required
def lista_productos(request):
    """Vista para listar productos con filtros avanzados"""
    from django.db.models import Sum, Prefetch, F
    
    # Query base con optimizaciones
    productos = Producto.objects.select_related('categoria').prefetch_related(
        Prefetch('stocks', queryset=Stock.objects.select_related('area'))
    ).filter(activo=True)
    
    # Agregar stock total y áreas con stock
    productos = productos.annotate(
        stock_total=Sum('stocks__cantidad'),
    )

    # FILTROS
    
    # 1. Filtro por búsqueda (nombre o código)
    busqueda = request.GET.get('q', '').strip()
    if busqueda:
        productos = productos.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )

    # 2. Filtro por categoría
    categoria_param = request.GET.get('categoria')
    categoria_seleccionada = None
    if categoria_param:
        try:
            categoria_id = int(categoria_param)
            productos = productos.filter(categoria_id=categoria_id)
            categoria_seleccionada = categoria_id
        except ValueError:
            pass

    # 3. Filtro por área
    area_param = request.GET.get('area')
    area_seleccionada = None
    if area_param:
        try:
            area_id = int(area_param)
            productos = productos.filter(stocks__area_id=area_id).distinct()
            area_seleccionada = area_id
        except ValueError:
            pass

    # 4. Filtro por estado de stock
    stock_filtro = request.GET.get('stock')
    if stock_filtro:
        if stock_filtro == 'agotado':
            productos = productos.filter(stock_total__lte=0)
        elif stock_filtro == 'bajo':
            productos = productos.filter(stock_total__lte=F('stock_minimo'), stock_total__gt=0)
        elif stock_filtro == 'disponible':
            productos = productos.filter(stock_total__gt=0)

    # ORDENAMIENTO
    orden = request.GET.get('orden', 'categoria__nombre')
    orden_valido = [
        'nombre', '-nombre', 'categoria__nombre', 'codigo', 
        '-fecha_creacion', 'stock_total', '-stock_total'
    ]
    
    if orden in orden_valido:
        productos = productos.order_by(orden, 'nombre')
    else:
        productos = productos.order_by('categoria__nombre', 'nombre')

    # Paginación
    from django.core.paginator import Paginator
    
    paginator = Paginator(productos, 12)  # 12 productos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Agregar información de áreas con stock y recalcular stock total para evitar valores desactualizados
    productos_con_areas = []
    for producto in page_obj:
        # Recalcular el stock total desde la base de datos (evita depender solo de la anotación)
        try:
            producto.stock_total = producto.stock_total()
        except TypeError:
            # Si existe colisión de nombre entre método y anotación, garantizamos el valor correcto
            producto.stock_total = producto.__class__.objects.filter(id=producto.id).aggregate(
                total=Sum('stocks__cantidad')
            )['total'] or 0

        areas_con_stock = producto.stocks.filter(cantidad__gt=0).values('area__nombre', 'cantidad')
        producto.areas_con_stock = areas_con_stock
        productos_con_areas.append(producto)

    # Datos para el template
    areas = Area.objects.filter(activo=True).order_by('nombre')
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')

    context = {
        'productos': productos_con_areas,
        'page_obj': page_obj,
        'areas': areas,
        'categorias': categorias,
        'area_seleccionada': area_seleccionada,
        'categoria_seleccionada': categoria_seleccionada,
        'busqueda': busqueda,
        'stock_filtro': stock_filtro,
        'orden_actual': orden,
        'total_productos': paginator.count,
    }

    return render(request, 'inventario/productos_con_guia.html', context)


@login_required
def detalle_producto(request, producto_id):
    """Vista de detalle de un producto"""
    producto = get_object_or_404(Producto, id=producto_id, activo=True)
    stocks = Stock.objects.select_related('area').filter(producto=producto)
    movimientos = Movimiento.objects.select_related(
        'area_origen', 'area_destino', 'usuario'
    ).filter(producto=producto).order_by('-fecha')[:20]
    
    context = {
        'producto': producto,
        'stocks': stocks,
        'movimientos': movimientos,
        'stock_total': producto.stock_total(),
        'tiene_stock_bajo': producto.tiene_stock_bajo(),
    }
    
    return render(request, 'inventario/detalle_producto.html', context)


@login_required
def alertas_stock(request):
    """Vista para mostrar alertas de stock"""
    alertas_list = []
    for alerta in AlertaStock.objects.select_related('producto', 'area').filter(estado='ACTIVA'):
        porcentaje = (alerta.stock_actual / alerta.stock_minimo * 100) if alerta.stock_minimo > 0 else 0
        critico = porcentaje < 50
        alertas_list.append({
            'alerta': alerta,
            'porcentaje': porcentaje,
            'critico': critico,
        })
    # Ordenar: críticos primero, luego por fecha descendente
    alertas_list = sorted(alertas_list, key=lambda x: (0 if x['critico'] else 1, x['alerta'].fecha_creacion), reverse=True)

    context = {
        'alertas': alertas_list,
    }

    return render(request, 'inventario/alertas.html', context)


def ayuda(request):
    """Vista de ayuda del sistema"""
    return render(request, 'inventario/ayuda.html')


@login_required
def api_stats(request):
    """API para estadísticas del dashboard"""
    # Datos para gráficos
    stock_por_categoria = []
    for categoria in Categoria.objects.filter(activo=True):
        total_stock = Stock.objects.filter(
            producto__categoria=categoria
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        
        stock_por_categoria.append({
            'categoria': categoria.nombre,
            'total': float(total_stock)
        })
    
    movimientos_recientes = []
    for movimiento in Movimiento.objects.order_by('-fecha')[:7]:
        movimientos_recientes.append({
            'fecha': movimiento.fecha.strftime('%Y-%m-%d'),
            'cantidad': float(movimiento.cantidad),
            'tipo': movimiento.tipo
        })
    
    return JsonResponse({
        'stock_por_categoria': stock_por_categoria,
        'movimientos_recientes': movimientos_recientes
    })


@ensure_csrf_cookie
@login_required
def agregar_producto(request):
    """Vista para agregar un nuevo producto con stock inicial"""
    if request.method == 'POST':
        form = AgregarProductoForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Crear el producto
                    producto = form.save()
                    
                    # Crear stock inicial si se proporcionó
                    area_inicial = form.cleaned_data.get('area_inicial')
                    cantidad_inicial = form.cleaned_data.get('cantidad_inicial')
                    
                    if area_inicial and cantidad_inicial and cantidad_inicial > 0:
                        # Crear entrada de stock inicial
                        numero_entrada = form.cleaned_data.get('numero_factura') or f'INICIAL-{producto.codigo}'
                        
                        entrada = EntradaStock.objects.create(
                            numero_entrada=numero_entrada,
                            tipo='COMPRA',
                            proveedor=form.cleaned_data.get('proveedor'),
                            fecha_compra=form.cleaned_data.get('fecha_compra') or date.today(),
                            total_compra=cantidad_inicial * (producto.precio_unitario or 0),
                            comprobante=form.cleaned_data.get('comprobante'),
                            observaciones=f'Stock inicial para producto {producto.nombre}',
                            registrado_por=request.user
                        )
                        
                        # Crear detalle de entrada
                        DetalleEntradaStock.objects.create(
                            entrada=entrada,
                            producto=producto,
                            cantidad=cantidad_inicial,
                            precio_unitario=producto.precio_unitario,
                            area_destino=area_inicial
                        )
                        
                        # Crear stock
                        Stock.objects.create(
                            producto=producto,
                            area=area_inicial,
                            cantidad=cantidad_inicial
                        )
                        
                        # Crear movimiento
                        Movimiento.objects.create(
                            producto=producto,
                            area_destino=area_inicial,
                            tipo='ENTRADA',
                            motivo='INICIAL',
                            cantidad=cantidad_inicial,
                            precio_unitario=producto.precio_unitario,
                            usuario=request.user,
                            observaciones=f'Stock inicial: {entrada.numero_entrada}'
                        )
                        
                        messages.success(request, f'¡Producto "{producto.nombre}" creado exitosamente con {cantidad_inicial} {producto.unidad_medida} en {area_inicial.nombre}!')
                    else:
                        messages.success(request, f'¡Producto "{producto.nombre}" creado exitosamente! Puedes agregar stock usando "Entrada de Stock".')
                    
                    return redirect('inventario:productos')
                    
            except Exception as e:
                messages.error(request, f'Error al crear el producto: {str(e)}')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = AgregarProductoForm()
    
    context = {
        'form': form,
        'categorias': Categoria.objects.filter(activo=True),
        'areas': Area.objects.filter(activo=True),
        'proveedores': Proveedor.objects.filter(activo=True),
        'area_tipos': Area.TIPOS_AREA,
    }
    return render(request, 'inventario/agregar_producto.html', context)


@login_required
def agregar_stock(request, producto_id):
    """Vista para agregar stock inicial a un producto recién creado"""
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        entrada_form = EntradaStockForm(request.POST, request.FILES)
        detalle_formset = DetalleEntradaFormSet(request.POST, prefix='detalles')
        
        if entrada_form.is_valid() and detalle_formset.is_valid():
            try:
                with transaction.atomic():
                    # Crear la entrada
                    entrada = entrada_form.save(commit=False)
                    entrada.registrado_por = request.user
                    entrada.save()
                    
                    # Procesar los detalles
                    for detalle_form in detalle_formset:
                        if detalle_form.cleaned_data and not detalle_form.cleaned_data.get('DELETE'):
                            detalle = detalle_form.save(commit=False)
                            detalle.entrada = entrada
                            detalle.save()
                            
                            # Actualizar o crear stock
                            stock, created = Stock.objects.get_or_create(
                                producto=detalle.producto,
                                area=detalle.area_destino,
                                defaults={'cantidad': detalle.cantidad}
                            )
                            if not created:
                                stock.cantidad += detalle.cantidad
                                stock.save()
                            
                            # Crear movimiento (entrada por compra)
                            Movimiento.objects.create(
                                producto=detalle.producto,
                                area_destino=detalle.area_destino,
                                tipo='ENTRADA',
                                motivo='COMPRA',
                                cantidad=detalle.cantidad,
                                precio_unitario=detalle.precio_unitario,
                                usuario=request.user,
                                observaciones=f'Entrada: {entrada.numero_entrada}'
                            )
                    
                    messages.success(request, f'¡Stock agregado exitosamente! Entrada registrada: {entrada.numero_entrada}')
                    return redirect('inventario:productos')
                    
            except Exception as e:
                messages.error(request, f'Error al procesar la entrada: {str(e)}')
    else:
        # Inicializar formularios
        entrada_form = EntradaStockForm()
        detalle_formset = DetalleEntradaFormSet(prefix='detalles')
        
        # Pre-llenar el primer detalle con el producto actual
        if detalle_formset.forms:
            detalle_formset.forms[0].initial = {'producto': producto}
    
    context = {
        'producto': producto,
        'entrada_form': entrada_form,
        'detalle_formset': detalle_formset,
        'proveedores': Proveedor.objects.filter(activo=True),
        'areas': Area.objects.filter(activo=True),
        'productos': Producto.objects.filter(activo=True),
    }
    return render(request, 'inventario/agregar_stock.html', context)


@login_required
def agregar_proveedor(request):
    """Vista para agregar un nuevo proveedor"""
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, f'¡Proveedor "{proveedor.nombre}" agregado exitosamente!')
            return redirect('inventario:agregar_stock', producto_id=request.GET.get('producto_id', 1))
    else:
        form = ProveedorForm()
    
    context = {
        'form': form,
        'producto_id': request.GET.get('producto_id'),
    }
    return render(request, 'inventario/agregar_proveedor.html', context)


@login_required
@ensure_csrf_cookie
def ingresar_factura(request):
    """Registrar una factura/boleta con uno o más productos (detalle múltiple)."""
    from datetime import date, datetime

    # Autogenerar número si el usuario lo deja vacío (opcional)
    def generar_numero():
        ultima = EntradaStock.objects.order_by('-id').first()
        if ultima and ultima.numero_entrada and ultima.numero_entrada.startswith('FAC-'):
            try:
                n = int(ultima.numero_entrada.split('-')[-1])
                return f"FAC-{n+1:04d}"
            except Exception:
                pass
        return f"FAC-{datetime.now().strftime('%m%d%H%M')}"

    if request.method == 'POST':
        entrada_form = EntradaStockForm(request.POST, request.FILES)
        detalle_formset = DetalleEntradaFormSet(request.POST, prefix='detalles')

        if entrada_form.is_valid() and detalle_formset.is_valid():
            try:
                with transaction.atomic():
                    entrada: EntradaStock = entrada_form.save(commit=False)
                    # Si no viene número, autogenerar
                    if not entrada.numero_entrada:
                        entrada.numero_entrada = generar_numero()
                    entrada.registrado_por = request.user
                    entrada.save()

                    # Procesar detalles
                    for form in detalle_formset:
                        if form.cleaned_data and not form.cleaned_data.get('DELETE'):
                            detalle = form.save(commit=False)
                            detalle.entrada = entrada
                            detalle.save()

                            # Actualizar o crear stock
                            stock, created = Stock.objects.get_or_create(
                                producto=detalle.producto,
                                area=detalle.area_destino,
                                defaults={'cantidad': detalle.cantidad}
                            )
                            if not created:
                                stock.cantidad += detalle.cantidad
                                stock.save()

                            # Crear movimiento enlazado a la factura
                            mov = Movimiento.objects.create(
                                producto=detalle.producto,
                                area_destino=detalle.area_destino,
                                tipo='ENTRADA',
                                motivo='COMPRA',
                                cantidad=detalle.cantidad,
                                precio_unitario=detalle.precio_unitario,
                                usuario=request.user,
                                observaciones=f'Entrada factura: {entrada.numero_entrada}'
                            )
                            # Enlaces de trazabilidad
                            mov.entrada = entrada
                            mov.detalle_entrada = detalle
                            mov.save(update_fields=['entrada', 'detalle_entrada'])

                    messages.success(request, f"Factura registrada: {entrada.numero_entrada} con {detalle_formset.total_form_count()} producto(s).")
                    return redirect('inventario:productos')
            except Exception as e:
                messages.error(request, f'Error al registrar la factura: {str(e)}')
        else:
            messages.error(request, 'Corrige los errores en el formulario.')
    else:
        entrada_form = EntradaStockForm(initial={'tipo': 'COMPRA', 'fecha_compra': date.today()})
        detalle_formset = DetalleEntradaFormSet(prefix='detalles')

    context = {
        'entrada_form': entrada_form,
        'detalle_formset': detalle_formset,
        'areas': Area.objects.filter(activo=True),
        'proveedores': Proveedor.objects.filter(activo=True),
        'productos': Producto.objects.filter(activo=True),
    }
    return render(request, 'inventario/ingresar_factura.html', context)


@login_required
def proveedores_sugeridos(request):
    """Devuelve proveedores sugeridos para un producto (+/- área) según historial."""
    producto_id = request.GET.get('producto_id')
    area_id = request.GET.get('area_id')
    if not producto_id:
        return JsonResponse({'success': False, 'error': 'producto_id requerido'}, status=400)

    detalles = DetalleEntradaStock.objects.filter(producto_id=producto_id).select_related('entrada__proveedor')
    if area_id:
        detalles = detalles.filter(area_destino_id=area_id)

    # Contar ocurrencias por proveedor
    from collections import Counter
    conteo = Counter()
    ultimo_por_prov = {}
    for d in detalles:
        prov = d.entrada.proveedor
        if prov:
            conteo[prov.id] += 1
            # Guardar última fecha/precio
            if (prov.id not in ultimo_por_prov) or (d.entrada.fecha_compra and d.entrada.fecha_compra > ultimo_por_prov[prov.id]['fecha']):
                ultimo_por_prov[prov.id] = {
                    'fecha': d.entrada.fecha_compra,
                    'precio': d.precio_unitario
                }

    sugeridos = []
    for prov_id, count in conteo.most_common(5):
        try:
            prov = Proveedor.objects.get(id=prov_id)
            info = ultimo_por_prov.get(prov_id, {})
            sugeridos.append({
                'id': prov.id,
                'nombre': prov.nombre,
                'veces': count,
                'ultima_fecha': info.get('fecha').isoformat() if info.get('fecha') else None,
                'ultimo_precio': str(info.get('precio')) if info.get('precio') is not None else None
            })
        except Proveedor.DoesNotExist:
            continue

    return JsonResponse({'success': True, 'sugeridos': sugeridos})


@login_required
def trazabilidad(request):
    """Reporte simple de trazabilidad por producto (timeline de movimientos)."""
    producto_id = request.GET.get('producto')
    movimientos = []
    producto_sel = None
    if producto_id:
        producto_sel = get_object_or_404(Producto, id=producto_id)
        movimientos = Movimiento.objects.select_related('area_origen','area_destino','usuario','entrada','detalle_entrada').filter(
            producto=producto_sel
        ).order_by('-fecha')[:200]

    return render(request, 'inventario/trazabilidad.html', {
        'productos': Producto.objects.filter(activo=True).order_by('nombre'),
        'producto_sel': producto_sel,
        'movimientos': movimientos,
    })


@login_required
@ensure_csrf_cookie
def entrada_stock(request):
    """Vista mejorada para registrar nueva entrada de stock con búsqueda de productos"""
    from datetime import date
    
    # Obtener parámetros de búsqueda
    busqueda = request.GET.get('q', '').strip()
    categoria_id = request.GET.get('categoria', '')
    area_id = request.GET.get('area', '')
    producto_param = request.GET.get('producto')
    
    # Generar próximo número de entrada automáticamente
    ultima_entrada_num = EntradaStock.objects.all().order_by('-id').first()
    if ultima_entrada_num:
        try:
            # Extraer número de la última entrada si tiene formato numérico
            ultimo_num = int(ultima_entrada_num.numero_entrada.split('-')[-1])
            proximo_numero = f"ENT-{ultimo_num + 1:04d}"
        except:
            # Si no se puede extraer número, usar timestamp
            from datetime import datetime
            proximo_numero = f"ENT-{datetime.now().strftime('%m%d%H%M')}"
    else:
        proximo_numero = "ENT-0001"
    
    # Construir queryset con filtros
    productos_query = Producto.objects.filter(activo=True).select_related('categoria')
    
    # Anotar con stock total
    productos_query = productos_query.annotate(
        stock_total=Sum('stocks__cantidad')
    )
    
    # Filtrar por búsqueda
    if busqueda:
        productos_query = productos_query.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    
    # Filtrar por categoría
    if categoria_id:
        productos_query = productos_query.filter(categoria_id=categoria_id)
    
    # Filtrar por área (productos que tengan stock en esa área)
    if area_id:
        productos_query = productos_query.filter(stocks__area_id=area_id).distinct()

    # Si viene un parámetro de producto específico, filtrar por ese id
    producto_seleccionada = None
    if producto_param:
        try:
            pid = int(producto_param)
            productos_query = productos_query.filter(id=pid)
            producto_seleccionada = pid
        except ValueError:
            pass
    
    # Obtener áreas con stock para cada producto y datos inteligentes
    productos = []
    for producto in productos_query[:50]:  # Limitar a 50 para mejor rendimiento
        areas_con_stock = Stock.objects.filter(
            producto=producto,
            cantidad__gt=0
        ).select_related('area')
        
        # Obtener datos de la última entrada de este producto para auto-completar
        ultima_entrada = DetalleEntradaStock.objects.filter(
            producto=producto
        ).select_related('entrada', 'entrada__proveedor', 'area_destino').order_by('-entrada__fecha_compra').first()
        
        # Agregar datos inteligentes al producto
        if ultima_entrada:
            producto.ultimo_proveedor = ultima_entrada.entrada.proveedor
            producto.ultima_area_destino = ultima_entrada.area_destino
            producto.ultimo_precio_unitario = ultima_entrada.precio_unitario
            producto.ultima_cantidad = ultima_entrada.cantidad
        else:
            producto.ultimo_proveedor = None
            producto.ultima_area_destino = None
            producto.ultimo_precio_unitario = None
            producto.ultima_cantidad = None
        
        producto.areas_con_stock = areas_con_stock
        productos.append(producto)
    
    # Manejo del POST para crear entrada de stock
    if request.method == 'POST':
        producto_id = request.POST.get('producto_id')
        numero_entrada = request.POST.get('numero_entrada')
        fecha_compra = request.POST.get('fecha_compra')
        proveedor_id = request.POST.get('proveedor')
        total_compra = request.POST.get('total_compra')
        area_destino_id = request.POST.get('area_destino')
        cantidad = request.POST.get('cantidad')
        precio_unitario = request.POST.get('precio_unitario')
        observaciones = request.POST.get('observaciones', '')
        comprobante = request.FILES.get('comprobante')
        
        # Identificación opcional: no exigimos numero_entrada
        if not all([producto_id, fecha_compra, area_destino_id, cantidad]):
            messages.error(request, 'Por favor completa todos los campos obligatorios.')
        else:
            try:
                with transaction.atomic():
                    # Crear la entrada
                    entrada = EntradaStock.objects.create(
                        numero_entrada=numero_entrada or '',
                        fecha_compra=fecha_compra,
                        proveedor_id=proveedor_id if proveedor_id else None,
                        total_compra=Decimal(total_compra) if total_compra else None,
                        observaciones=observaciones,
                        comprobante=comprobante,
                        registrado_por=request.user
                    )
                    
                    # Obtener producto y área
                    producto = Producto.objects.get(id=producto_id)
                    area = Area.objects.get(id=area_destino_id)
                    cantidad_decimal = Decimal(cantidad)
                    precio_decimal = Decimal(precio_unitario) if precio_unitario else None
                    
                    # Crear detalle de entrada
                    detalle = DetalleEntradaStock.objects.create(
                        entrada=entrada,
                        producto=producto,
                        cantidad=cantidad_decimal,
                        precio_unitario=precio_decimal,
                        area_destino=area
                    )
                    
                    # Actualizar o crear stock
                    stock, created = Stock.objects.get_or_create(
                        producto=producto,
                        area=area,
                        defaults={'cantidad': cantidad_decimal}
                    )
                    if not created:
                        stock.cantidad += cantidad_decimal
                        stock.save()
                    
                    # Crear movimiento (entrada por compra)
                    Movimiento.objects.create(
                        producto=producto,
                        area_destino=area,
                        tipo='ENTRADA',
                        motivo='COMPRA',
                        cantidad=cantidad_decimal,
                        precio_unitario=precio_decimal,
                        usuario=request.user,
                        observaciones=f'Entrada: {entrada.numero_entrada}'
                    )
                    
                    messages.success(request, f'¡Stock agregado exitosamente! Entrada registrada: {entrada.numero_entrada}')
                    # Redirigir manteniendo filtros
                    url_params = {}
                    if busqueda:
                        url_params['q'] = busqueda
                    if categoria_id:
                        url_params['categoria'] = categoria_id
                    if area_id:
                        url_params['area'] = area_id
                    
                    if url_params:
                        return redirect(f"{reverse('inventario:entrada_stock')}?{urlencode(url_params)}")
                    else:
                        return redirect('inventario:entrada_stock')
                        
            except Exception as e:
                messages.error(request, f'Error al procesar la entrada: {str(e)}')
    
    context = {
        'productos': productos,
        'total_productos': len(productos),
        'busqueda': busqueda,
        'categoria_seleccionada': int(categoria_id) if categoria_id else None,
        'area_seleccionada': int(area_id) if area_id else None,
        'categorias': Categoria.objects.filter(activo=True),
        'areas': Area.objects.filter(activo=True),
        'proveedores': Proveedor.objects.filter(activo=True),
        'fecha_hoy': date.today().isoformat(),
        'proximo_numero_entrada': proximo_numero,
        'producto_seleccionada': producto_seleccionada,
    }
    return render(request, 'inventario/entrada_stock.html', context)


@login_required
def reporte_productos_stock(request):
    """Genera un reporte de productos con stock en formato CSV o PDF."""
    from django.db.models import Sum
    import openpyxl
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from io import BytesIO

    formato = request.GET.get('formato', 'csv').lower()

    productos = Producto.objects.filter(activo=True).annotate(stock_total=Sum('stocks__cantidad')).filter(stock_total__gt=0)

    if formato == 'pdf':
        # Generar PDF con formato optimizado para legibilidad
        from reportlab.lib.pagesizes import letter, A4, landscape
        from reportlab.platypus import PageBreak, LongTable
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

        buffer = BytesIO()
        # Usar landscape para mejor ajuste
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []

        styles = getSampleStyleSheet()

        # Título con mejor formato
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        title = Paragraph("Reporte de Productos con Stock", title_style)
        elements.append(title)

        # Información del reporte
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        info_text = f"Total de productos: {productos.count()} | Generado el {date.today().strftime('%d/%m/%Y')}"
        elements.append(Paragraph(info_text, info_style))

        # Datos para la tabla optimizada
        data = [['ID', 'Código', 'Nombre del Producto', 'Categoría', 'Unidad', 'Stock', 'Precio', 'Áreas']]

        for p in productos:
            areas = Stock.objects.filter(producto=p, cantidad__gt=0).select_related('area')
            areas_info = ', '.join([f"{a.area.nombre}({a.cantidad})" for a in areas])

            # Limitar longitudes para mejor ajuste
            nombre_corto = p.nombre[:25] + '...' if len(p.nombre) > 25 else p.nombre
            categoria_corta = (p.categoria.nombre if p.categoria else 'Sin cat.')[:15] + '...' if p.categoria and len(p.categoria.nombre) > 15 else (p.categoria.nombre if p.categoria else 'Sin cat.')
            areas_corto = areas_info[:35] + '...' if len(areas_info) > 35 else areas_info

            data.append([
                str(p.id),
                p.codigo or '-',
                nombre_corto,
                categoria_corta,
                p.get_unidad_medida_display() if hasattr(p, 'get_unidad_medida_display') else str(p.unidad_medida or '-'),
                str(p.stock_total or 0),
                f"${p.precio_unitario:.2f}" if p.precio_unitario else '-',
                areas_corto
            ])

        # Anchos de columna optimizados para landscape
        col_widths = [25, 50, 120, 70, 45, 40, 55, 120]  # Más compacto

        # Usar LongTable para múltiples páginas si es necesario
        table = LongTable(data, colWidths=col_widths, repeatRows=1)

        # Estilos optimizados para mejor legibilidad
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),

            # Cuerpo de la tabla
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('TOPPADDING', (0, 1), (-1, -1), 3),

            # Alineación específica por columna
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # ID centrado
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),  # Código centrado
            ('ALIGN', (4, 0), (4, -1), 'CENTER'),  # Unidad centrada
            ('ALIGN', (5, 0), (5, -1), 'CENTER'),  # Stock centrado
            ('ALIGN', (6, 0), (6, -1), 'RIGHT'),   # Precio alineado a la derecha

            # Bordes sutiles
            ('GRID', (0, 0), (-1, -1), 0.3, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),

            # Filas alternas para mejor legibilidad (más sutil)
            ('BACKGROUND', (0, 2), (-1, 2), colors.lightgrey),
            ('BACKGROUND', (0, 4), (-1, 4), colors.lightgrey),
            ('BACKGROUND', (0, 6), (-1, 6), colors.lightgrey),
            ('BACKGROUND', (0, 8), (-1, 8), colors.lightgrey),
        ]))

        elements.append(table)

        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="productos_con_stock.pdf"'
        return response

    elif formato == 'excel':
        # Generar Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos con Stock"

        # Encabezados
        headers = ['ID', 'Código', 'Nombre', 'Categoría', 'Unidad', 'Stock Total', 'Precio Unitario', 'Áreas (nombre:cantidad)']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Datos
        for row_num, p in enumerate(productos, 2):
            areas = Stock.objects.filter(producto=p, cantidad__gt=0).select_related('area')
            areas_info = '; '.join([f"{a.area.nombre}:{a.cantidad}" for a in areas])

            ws.cell(row=row_num, column=1, value=p.id)
            ws.cell(row=row_num, column=2, value=p.codigo)
            ws.cell(row=row_num, column=3, value=p.nombre)
            ws.cell(row=row_num, column=4, value=p.categoria.nombre if p.categoria else '')
            ws.cell(row=row_num, column=5, value=p.get_unidad_medida_display() if hasattr(p, 'get_unidad_medida_display') else p.unidad_medida)
            ws.cell(row=row_num, column=6, value=p.stock_total or 0)
            ws.cell(row=row_num, column=7, value=str(p.precio_unitario or ''))
            ws.cell(row=row_num, column=8, value=areas_info)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="productos_con_stock.xlsx"'
        return response

    else:
        # Generar CSV (por defecto)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="productos_con_stock.csv"'

        writer = csv.writer(response)
        writer.writerow(['ID', 'Código', 'Nombre', 'Categoría', 'Unidad', 'Stock Total', 'Precio Unitario', 'Áreas (nombre:cantidad)'])

        for p in productos:
            areas = Stock.objects.filter(producto=p, cantidad__gt=0).select_related('area')
            areas_info = '; '.join([f"{a.area.nombre}:{a.cantidad}" for a in areas])
            writer.writerow([
                p.id,
                p.codigo,
                p.nombre,
                p.categoria.nombre if p.categoria else '',
                p.get_unidad_medida_display() if hasattr(p, 'get_unidad_medida_display') else p.unidad_medida,
                str(p.stock_total or 0),
                str(p.precio_unitario or ''),
                areas_info
            ])

        return response


@login_required
@ensure_csrf_cookie
def salida_stock(request):
    """Registrar una salida de stock (consumo/uso) con búsqueda y filtros avanzados."""
    from datetime import date
    from decimal import Decimal

    # Obtener parámetros de búsqueda y filtros
    busqueda = request.GET.get('q', '').strip()
    categoria_id = request.GET.get('categoria', '')
    area_id = request.GET.get('area', '')
    producto_param = request.GET.get('producto')

    # Construir queryset con filtros y solo productos con stock
    productos_query = Producto.objects.filter(activo=True).select_related('categoria')

    # Anotar con stock total y filtrar solo productos con stock > 0
    productos_query = productos_query.annotate(
        stock_total=Sum('stocks__cantidad')
    ).filter(stock_total__gt=0)

    # Filtrar por búsqueda
    if busqueda:
        productos_query = productos_query.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )

    # Filtrar por categoría
    if categoria_id:
        productos_query = productos_query.filter(categoria_id=categoria_id)

    # Filtrar por área (productos que tengan stock en esa área)
    if area_id:
        productos_query = productos_query.filter(stocks__area_id=area_id).distinct()

    # Si viene un parámetro de producto específico, filtrar por ese id
    producto_seleccionada = None
    if producto_param:
        try:
            pid = int(producto_param)
            productos_query = productos_query.filter(id=pid)
            producto_seleccionada = pid
        except ValueError:
            pass

    # Obtener áreas con stock para cada producto
    productos = []
    for producto in productos_query[:50]:  # Limitar a 50 para mejor rendimiento
        areas_con_stock = Stock.objects.filter(
            producto=producto,
            cantidad__gt=0
        ).select_related('area').order_by('area__nombre')

        producto.areas_con_stock = areas_con_stock
        productos.append(producto)

    if request.method == 'POST':
        producto_id = request.POST.get('producto')
        area_id = request.POST.get('area')
        cantidad = request.POST.get('cantidad')
        motivo = request.POST.get('motivo') or 'CONSUMO'

        if not producto_id or not area_id or not cantidad:
            messages.error(request, 'Completa los campos obligatorios.')
            return redirect('inventario:salida_stock')

        try:
            producto = Producto.objects.get(id=producto_id)
            area = Area.objects.get(id=area_id)
            cantidad_dec = Decimal(str(cantidad))
            if cantidad_dec <= 0:
                messages.error(request, 'Cantidad debe ser mayor a 0.')
                return redirect('inventario:salida_stock')

            stock = Stock.objects.filter(producto=producto, area=area).first()
            if not stock or stock.cantidad < cantidad_dec:
                messages.error(request, 'No hay suficiente stock en el área seleccionada.')
                return redirect('inventario:salida_stock')

            with transaction.atomic():
                stock.cantidad -= cantidad_dec
                stock.save()

                Movimiento.objects.create(
                    producto=producto,
                    area_origen=area,
                    area_destino=None,
                    tipo='SALIDA',
                    motivo=motivo,
                    cantidad=cantidad_dec,
                    precio_unitario=None,
                    usuario=request.user,
                    observaciones=f'Salida por {motivo}'
                )

            messages.success(request, f'Salida registrada: -{cantidad_dec} {producto.unidad_medida} de {producto.nombre} en {area.nombre}.')
            return redirect('inventario:detalle_producto', producto_id=producto.id)

        except Producto.DoesNotExist:
            messages.error(request, 'Producto no encontrado.')
            return redirect('inventario:salida_stock')
        except Area.DoesNotExist:
            messages.error(request, 'Área no encontrada.')
            return redirect('inventario:salida_stock')
        except Exception as e:
            messages.error(request, f'Error interno: {str(e)}')
            return redirect('inventario:salida_stock')

    # GET: mostrar formulario con búsqueda y filtros
    context = {
        'productos': productos,
        'total_productos': len(productos),
        'busqueda': busqueda,
        'categoria_seleccionada': int(categoria_id) if categoria_id else None,
        'area_seleccionada': int(area_id) if area_id else None,
        'categorias': Categoria.objects.filter(activo=True),
        'areas': Area.objects.filter(activo=True),
        'producto_seleccionada': producto_seleccionada,
    }
    return render(request, 'inventario/salida_stock.html', context)


@login_required
@ensure_csrf_cookie
def entrada_stock_simple(request):
    """Vista simplificada para entrada de stock con interfaz de cards"""
    from datetime import date
    
    # Generar próximo número de entrada automáticamente
    ultima_entrada_num = EntradaStock.objects.all().order_by('-id').first()
    if ultima_entrada_num:
        try:
            ultimo_num = int(ultima_entrada_num.numero_entrada.split('-')[-1])
            proximo_numero = f"REC-{ultimo_num + 1:04d}"
        except:
            from datetime import datetime
            proximo_numero = f"REC-{datetime.now().strftime('%m%d%H%M')}"
    else:
        proximo_numero = "REC-0001"
    
    # Manejo del POST: crear entrada de stock simple
    if request.method == 'POST':
        try:
            with transaction.atomic():
                producto_id = request.POST.get('producto_id') or request.POST.get('producto') or request.POST.get('idProducto')
                cantidad = request.POST.get('cantidad')
                area_id = request.POST.get('area') or request.POST.get('area_destino')
                proveedor_id = request.POST.get('proveedor')
                precio = request.POST.get('precio') or request.POST.get('precio_unitario')
                fecha_compra = request.POST.get('fecha_compra') or date.today().isoformat()

                if not producto_id or not cantidad or not area_id:
                    messages.error(request, 'Faltan datos obligatorios: producto, cantidad y área.')
                    return redirect('inventario:entrada_stock')

                # Generar número de recibo/entrada
                ultima_entrada_num = EntradaStock.objects.all().order_by('-id').first()
                if ultima_entrada_num:
                    try:
                        ultimo_num = int(ultima_entrada_num.numero_entrada.split('-')[-1])
                        numero_generado = f"REC-{ultimo_num + 1:04d}"
                    except Exception:
                        from datetime import datetime
                        numero_generado = f"REC-{datetime.now().strftime('%m%d%H%M%S')}"
                else:
                    numero_generado = "REC-0001"

                # Objetos base
                producto = Producto.objects.get(id=producto_id)
                area = Area.objects.get(id=area_id)
                cantidad_dec = Decimal(str(cantidad))
                if cantidad_dec <= 0:
                    messages.error(request, 'La cantidad debe ser mayor a 0.')
                    return redirect('inventario:entrada_stock')
                precio_dec = Decimal(str(precio)) if precio else None

                # Crear entrada y detalle
                entrada = EntradaStock.objects.create(
                    numero_entrada=numero_generado,
                    tipo='COMPRA',
                    proveedor_id=proveedor_id if proveedor_id else None,
                    fecha_compra=fecha_compra,
                    total_compra=(cantidad_dec * precio_dec) if precio_dec else None,
                    observaciones=f'Entrada rápida - {producto.nombre}',
                    registrado_por=request.user
                )

                DetalleEntradaStock.objects.create(
                    entrada=entrada,
                    producto=producto,
                    cantidad=cantidad_dec,
                    precio_unitario=precio_dec,
                    area_destino=area
                )

                # Actualizar/crear stock
                stock, created = Stock.objects.get_or_create(
                    producto=producto,
                    area=area,
                    defaults={'cantidad': cantidad_dec}
                )
                if not created:
                    stock.cantidad += cantidad_dec
                    stock.save()

                # Registrar movimiento
                Movimiento.objects.create(
                    producto=producto,
                    area_destino=area,
                    tipo='ENTRADA',
                    motivo='COMPRA',
                    cantidad=cantidad_dec,
                    precio_unitario=precio_dec,
                    usuario=request.user,
                    observaciones=f'Entrada: {entrada.numero_entrada}'
                )

                messages.success(request, f'Stock agregado: +{cantidad_dec} {producto.unidad_medida} a {producto.nombre}.')
                return redirect('inventario:detalle_producto', producto_id=producto.id)
        except Exception as e:
            messages.error(request, f'Error al registrar la entrada: {str(e)}')
            return redirect('inventario:entrada_stock')

    # Obtener productos con stock actual
    productos = Producto.objects.filter(activo=True).select_related('categoria').annotate(
        stock_total=Sum('stocks__cantidad')
    ).order_by('categoria__nombre', 'nombre')
    
    # Agregar información de última entrada para auto-completar
    productos_con_info = []
    for producto in productos:
        ultima_entrada = DetalleEntradaStock.objects.filter(
            producto=producto
        ).select_related('entrada', 'entrada__proveedor', 'area_destino').order_by('-entrada__fecha_compra').first()
        
        if ultima_entrada:
            producto.ultimo_proveedor = ultima_entrada.entrada.proveedor
            producto.ultima_area = ultima_entrada.area_destino
            producto.ultimo_precio = ultima_entrada.precio_unitario
        else:
            producto.ultimo_proveedor = None
            producto.ultima_area = None
            producto.ultimo_precio = producto.precio_unitario or 0
            
        productos_con_info.append(producto)
    
    context = {
        'productos': productos_con_info,
        'areas': Area.objects.filter(activo=True),
        'proveedores': Proveedor.objects.filter(activo=True),
        'categorias': Categoria.objects.filter(activo=True),
        'fecha_hoy': date.today().isoformat(),
        'proximo_numero': proximo_numero,
    }
    return render(request, 'inventario/entrada_stock_simple.html', context)


@login_required
def agregar_stock_ajax(request):
    """Vista AJAX para agregar stock de forma rápida"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            
            producto_id = data.get('producto_id')
            cantidad = Decimal(str(data.get('cantidad', 0)))
            area_id = data.get('area_id')
            proveedor_id = data.get('proveedor_id')
            precio_unitario = data.get('precio_unitario')
            numero_recibo = data.get('numero_recibo')
            
            if not all([producto_id, cantidad, area_id]):
                return JsonResponse({
                    'success': False,
                    'error': 'Faltan datos obligatorios'
                })
            
            if cantidad <= 0:
                return JsonResponse({
                    'success': False,
                    'error': 'La cantidad debe ser mayor a 0'
                })
            
            with transaction.atomic():
                # Obtener objetos
                producto = Producto.objects.get(id=producto_id)
                area = Area.objects.get(id=area_id)
                proveedor = Proveedor.objects.get(id=proveedor_id) if proveedor_id else None
                
                # Crear entrada de stock
                entrada = EntradaStock.objects.create(
                    numero_entrada=numero_recibo,
                    fecha_compra=date.today(),
                    proveedor=proveedor,
                    total_compra=cantidad * (Decimal(str(precio_unitario)) if precio_unitario else 0),
                    observaciones=f'Entrada rápida - {producto.nombre}',
                    registrado_por=request.user
                )
                
                # Crear detalle
                DetalleEntradaStock.objects.create(
                    entrada=entrada,
                    producto=producto,
                    cantidad=cantidad,
                    precio_unitario=Decimal(str(precio_unitario)) if precio_unitario else None,
                    area_destino=area
                )
                
                # Actualizar stock
                stock, created = Stock.objects.get_or_create(
                    producto=producto,
                    area=area,
                    defaults={'cantidad': cantidad}
                )
                if not created:
                    stock.cantidad += cantidad
                    stock.save()
                
                # Crear movimiento
                Movimiento.objects.create(
                    producto=producto,
                    area_destino=area,
                    tipo='ENTRADA',
                    motivo='COMPRA',
                    cantidad=cantidad,
                    precio_unitario=Decimal(str(precio_unitario)) if precio_unitario else None,
                    usuario=request.user,
                    observaciones=f'Entrada: {entrada.numero_entrada}'
                )
                
                # Calcular nuevo stock total
                nuevo_stock_total = Stock.objects.filter(producto=producto).aggregate(
                    total=Sum('cantidad')
                )['total'] or 0
                
                return JsonResponse({
                    'success': True,
                    'message': f'Stock agregado exitosamente',
                    'nuevo_stock': str(nuevo_stock_total),
                    'recibo': entrada.numero_entrada
                })
                
        except Producto.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Producto no encontrado'
            })
        except Area.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Área no encontrada'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error interno: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Método no permitido'
    })


def api_entities(request):
    """Devuelve listas de proveedores y áreas (para poblar selects).

    Si no está autenticado, devuelve JSON con error 401 (evita redirección HTML).
    """
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'No autenticado'}, status=401)

    proveedores = list(Proveedor.objects.filter(activo=True).values('id', 'nombre'))
    areas = list(Area.objects.filter(activo=True).values('id', 'nombre'))
    return JsonResponse({'success': True, 'proveedores': proveedores, 'areas': areas})


def api_create_entity(request):
    """Crea un proveedor o un área vía AJAX (espera JSON: {type:'proveedor'|'area', name: '...'}).

    Si el usuario no está autenticado devuelve JSON 401 en lugar de redirigir.
    """
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'No autenticado'}, status=401)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    try:
        data = json.loads(request.body)
        typ = data.get('type')
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Nombre requerido'}, status=400)

        if typ == 'proveedor':
            # Proveedor requiere rut único; generar uno automático si no se proporciona
            rut = data.get('rut') or f'AUTO-{int(time.time())}'
            prov = Proveedor.objects.create(nombre=name, rut=rut, activo=True)
            return JsonResponse({'success': True, 'id': prov.id, 'nombre': prov.nombre, 'type': 'proveedor'})

        elif typ == 'area':
            tipo = data.get('tipo') or 'BODEGA'
            area = Area.objects.create(nombre=name, tipo=tipo, activo=True)
            return JsonResponse({'success': True, 'id': area.id, 'nombre': area.nombre, 'type': 'area'})

        else:
            return JsonResponse({'success': False, 'error': 'Tipo inválido'}, status=400)

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def api_delete_entity(request):
    """Marca como inactivo un proveedor o área (espera JSON: {type, id}).

    Devuelve JSON 401 si no está autenticado para evitar redirecciones HTML.
    """
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'No autenticado'}, status=401)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    try:
        data = json.loads(request.body)
        typ = data.get('type')
        obj_id = data.get('id')
        if not obj_id:
            return JsonResponse({'success': False, 'error': 'ID requerido'}, status=400)

        if typ == 'proveedor':
            prov = get_object_or_404(Proveedor, id=obj_id)
            prov.activo = False
            prov.save()
            return JsonResponse({'success': True})

        elif typ == 'area':
            area = get_object_or_404(Area, id=obj_id)
            area.activo = False
            area.save()
            return JsonResponse({'success': True})

        else:
            return JsonResponse({'success': False, 'error': 'Tipo inválido'}, status=400)

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
